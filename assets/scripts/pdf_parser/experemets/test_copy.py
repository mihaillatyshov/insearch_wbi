import openpyxl

src_wb = openpyxl.load_workbook("C:/Coding/works/wbi/pdf_parser/drill_ALM/xlsx/res/sm_simple.xlsx")
dest_wb = openpyxl.Workbook()

src_sheet = src_wb.get_sheet_by_name('source_sheet')
dest_sheet = dest_wb.get_sheet_by_name('destination')

for i in range(1, src_sheet.max_row + 1):
    for j in range(1, src_sheet.max_column + 1):
        dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

src_wb.save('source.xlsx')
dest_wb.save('destination.xlsx')
