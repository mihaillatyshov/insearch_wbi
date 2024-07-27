import os

import openpyxl as op

from shared import ArgsBase, parse_args


class Args(ArgsBase):
    xlsx_path: str
    save_path: str


START_ROW = 2                                                                                                           #номер строки, с которой начинаются модели
WRITE_ROW = 2


def parse_xl(xlsx_path: os.DirEntry[str], heads: list[str], sm_ws):
    # for file in glob.glob(path_xl):
    global START_ROW
    global WRITE_ROW
    print(xlsx_path.name)
    xl = op.load_workbook(xlsx_path.path)
    ws = xl.active
    temp_cols = []                                                                                                      #временный список номеров колонок в СМ
    temp_heads = []                                                                                                     #временный список названий столбцов

    #собираем названия столбцов во временный список, добавляем новые столбцы в СМ
    k = 0
    for col in range(1, ws.max_column + 1):
        head = str(ws.cell(1, col).value)

        #если заголовок уже есть во временном списке, то добавляем индекс к заголовку
        if head in temp_heads:
            k += 1
            head = head + str(k)

        if head not in heads:
            heads.append(head)
            c = len(heads)                                                                                              #номер столбца в СМ
            temp_cols.append(c)
            sm_ws.cell(1, c).value = head                                                                               # записываем новый заголовок в эксель
        else:
            c = heads.index(head) + 1
            temp_cols.append(c)
        temp_heads.append(head)

    #собираем данные для каждой модели построчно
    for row in range(START_ROW, ws.max_row + 1):
        sm_ws.cell(WRITE_ROW, 1).value = xlsx_path.name
        for col1 in range(1, ws.max_column + 1):
            col_sm = temp_cols[col1 - 1]
            value = ws.cell(row, col1).value
            sm_ws.cell(WRITE_ROW, col_sm).value = value
        WRITE_ROW += 1


def create_sm(args: Args):
    res_xl = op.Workbook()
    sm_ws = res_xl.active
    sm_ws.title = "sm"
    heads = ['FILE']
    sm_ws.cell(1, 1).value = 'FILE'

    os.makedirs(args.save_path, exist_ok=True)

    for filename in os.scandir(args.xlsx_path):
        if filename.is_file():
            if (filename.name.startswith("~$")):
                continue
            parse_xl(filename, heads, sm_ws)

    res_xl.save(os.path.join(args.save_path, "2_sm_simple.xlsx"))


create_sm(parse_args(Args))

# python 2_create_sm.py C:/Coding/works/wbi/amati_drill/xlsx_add_info/ C:/Coding/works/wbi/amati_drill/xlsx_sm/
