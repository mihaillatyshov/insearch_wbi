import os

import openpyxl
from openpyxl.styles import Font, PatternFill
from shared import ArgsBase, openpyxl_ws_autofit, parse_args


class Args(ArgsBase):
    xlsx_path: str
    save_path: str


py_args: Args = parse_args(Args)

# path = "./20230724-122626-complete_dso.xlsx" #путь к исходному файлу
path = py_args.xlsx_path                                                                                                #путь к исходному файлу

xl_file = openpyxl.load_workbook(path)
ws_sm = xl_file['sm']
constr_str = "constr"

sheets_ex = xl_file.sheetnames
p_dso = 8
p_model = p_dso + 1
heads = [ws_sm.cell(1, i).value for i in range(1, ws_sm.max_column + 1)]
col_constr = heads.index(constr_str) + 1
max_row = ws_sm.max_row + 1
for row in range(2, max_row):
    print(row)
    constr = ws_sm.cell(row, col_constr).value
    if constr is None:
        continue
    elif constr not in sheets_ex:
        ws_sm.cell(row, 1).fill = PatternFill("solid", fgColor="00FFFF00")
        continue
    ws = xl_file[constr]
    c_dso = ws.max_column
    c_new_head = c_dso + 1                                                                                              #имена заголовков в dso

    heads_dso = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    col_m = heads_dso.index('model') + 1
    if ws.cell(9, col_m).value is None:
        p_model = 9
    else:
        p_model = ws.max_row + 1

    #имена заголовков из sm в dso
    heads_new_dso = [ws.cell(p_dso, i).value for i in range(1, ws.max_column + 1)]
    for col in range(2, ws_sm.max_column + 1):
        head = ws_sm.cell(1, col).value
        value = ws_sm.cell(row, col).value
        if value is None:
            continue
        if head == constr_str or head == 'model_1':
            continue
        if head in heads_dso:
            ind = heads_dso.index(head) + 1
            if ws.cell(p_dso, ind).value is None:
                ws.cell(p_dso, ind).value = head
            ws.cell(p_model, ind).value = value
        # если колонок нет в dso заполняем в дополнительнные колонки с заголовками, следущие за основными
        else:
            if head in heads_new_dso:
                c_new_head = heads_new_dso.index(head) + 1
            else:
                c_new_head = ws.max_column + 1
                ws.cell(p_dso, c_new_head).value = head
            ws.cell(p_model, c_new_head).value = value
# в каждом шаблоне дополнительные колонки с заголовками переносим в пустые колонки и отмечаем желтым цветом
for sheet in sheets_ex:
    print(sheet)
    if sheet == 'sm' or sheet == 'repr_fields':
        continue
    ws = xl_file[sheet]
    heads_dso1 = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

    # NOTE my fix
    if heads_dso1[-1] is not None:
        heads_dso1.append(None)

    col1 = heads_dso1.index(None) + 1
    heads_new_dso1 = [ws.cell(p_dso, i).value for i in range(1, ws.max_column + 1)]
    while ws.cell(8, col1).value is not None:
        ind = heads_new_dso1.index(None)
        empty_col = ind + 1
        heads_new_dso1[ind] = ws.cell(8, col1).value
        ws.cell(8, empty_col).value = ws.cell(8, col1).value
        ws.cell(8, empty_col).fill = PatternFill("solid", fgColor="00FFFF00")
        for row in range(9, ws.max_row + 1):
            value = ws.cell(row, col1).value
            ws.cell(row, empty_col).value = value
        col1 += 1

    openpyxl_ws_autofit(ws)

xl_file.save(os.path.join(py_args.save_path, '4_complete_dso.xlsx'))
xl_file.close()

   ## "./xlsx/multiple/2_db_export_from_toolz.xlsx"

   # * python ctd_from_sm_fill.py C:/Coding/works/wbi/amati_drill/xlsx_sm/20230921-140624-export_from_toolz.xlsx C:/Coding/works/wbi/amati_drill/xlsx_sm/
