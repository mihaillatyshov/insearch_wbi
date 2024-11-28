import argparse
import os
import time as t
from collections import namedtuple

import openpyxl
import pandas as pd
import pandas.io.sql as sql
import psycopg2
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from shared import get_onedrive_path

arg_parser = argparse.ArgumentParser(prog='3_yg1_load_db_ctd')
arg_parser.add_argument('path')
cmd_args = arg_parser.parse_args()

SM_SIMPLE_PATH = os.path.join(get_onedrive_path(), "Work/wbi/yg1-shop/parsing/out", cmd_args.path, "2_sm_simple.xlsx")
EXPORT_FOLDER = os.path.join(get_onedrive_path(), "Work/wbi/yg1-shop/parsing/out", cmd_args.path)

df = pd.read_excel(SM_SIMPLE_PATH, index_col=None, engine="openpyxl")
constr_set = set(df["constr"])
print("Constr to process: ", constr_set)


def open_connection():
    config = {
        'user': 'postgres',
        'password': 'Radius314',
        'host': '192.168.111.115',
        'port': '5432',
        'database': 'toolz370_01'
    }
    try:
        conn = psycopg2.connect(**config)
        cursor = conn.cursor()
    except:
        print('Не могу подключиться к базе')
    return (conn, cursor)


def get_db_rows(curs, sql, args=None):                                                                                  # sql как параметр
    curs.execute(sql, args)
    columns = [col[0] for col in curs.description]
    RowType = namedtuple('Row', columns)
    data = [RowType(*row) for row in curs.fetchall()]
    return (data)                                                                                                       # стандартный разбор Sql запроса.


timestr = t.strftime("%Y%m%d-%H%M%S")
dest_export_path = os.path.join(EXPORT_FOLDER, "3_export_from_toolz.xlsx")

xl_file_dso = openpyxl.Workbook()
dso_sheet = xl_file_dso.active
dso_sheet.title = 'first'
xl_file_dso.save(dest_export_path)
xl_file_dso.close()

## Подключаемся к БД
co, cu = open_connection()

# считываем таблицы, которые нужно получить в твёрдом xl
# Читать в лист из текстовика
tables_list = list(constr_set)

# Прописываем Writer, чтобы записать всё в один файл
writer = pd.ExcelWriter(dest_export_path, engine='openpyxl')
writer.workbook = xl_file_dso
writer.sheets.update(dict((ws.title, ws) for ws in xl_file_dso.worksheets))

# read and save the data
for table in tables_list:
    if table == '':
        continue
    try:
        query = "SELECT * FROM tools." + table + " ORDER BY random() LIMIT 5"
        df = sql.read_sql(query, co)
        df = df[0:0]
        df = df.drop(columns=['tool_id', 'item_id'])
        # export the data into the excel sheet
        df.to_excel(writer, sheet_name=table, index=False)

    except Exception as exc:
        print(f"Somethin wrong with table: '{table}'", f"Error: {exc}")
        continue

writer.close()
print('The previous file has been created')

# Некоторые дополнительные действия для экономии своих сил и времени.

dso_file_previous = openpyxl.load_workbook(dest_export_path)
# удаляем 1 лист
# std=dso_file_previous['first']
# dso_file_previous.remove(std)
# print('empty sheet removed.')

# удаляем все колонки ID во всех листах
for wb_name in dso_file_previous.sheetnames:
    wb = dso_file_previous[wb_name]
    wb.delete_cols(1, 1)

repr_query = ("SELECT * FROM pg_catalog.pg_tables where tablename like 'descr%';")
resu = get_db_rows(cu, repr_query)
orange_fields = list()
for ro in resu:
    field_name = str(ro.tablename).replace('descr_', '')
    orange_fields.append(field_name)

# теперь выделим все обязательные поля

clear_orange = list()

# Прописываем комментарии для каждого поля
for constr in dso_file_previous.sheetnames:
    wb = dso_file_previous[constr]
    for j in range(1, wb.max_column + 1):
        field = wb.cell(1, j).value
        query = ("SELECT * FROM tools.gen_fields gf WHERE fieldname = %s")
        res = get_db_rows(cu, query, (field, ))
        allow_nulls = 0
        if len(res) > 0:
            for row2 in res:
                RuDescr = row2.rudescription
                unitRu = row2.unitru
                allow_nulls = row2.allownulls
            if (unitRu is None):
                comment_ = Comment(RuDescr, "Alexander Vinogradov")
            else:
                comment_ = Comment(RuDescr + " (" + unitRu + ")", "Alexander Vinogradov")

            if allow_nulls == 0:
                wb.cell(1, j).fill = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")

            wb.cell(1, j).comment = comment_

        if field in orange_fields:
            wb.cell(1, j).fill = PatternFill(start_color="FD6A02", end_color="FD6A02", fill_type="solid")               # оранжевый
            tup = (field, RuDescr)
            clear_orange.append(tup)

orange_uniq = list(dict.fromkeys(clear_orange))
dso_file_previous.create_sheet('repr_fields')
ws = dso_file_previous["repr_fields"]
q = 1
for orange_field_p in orange_uniq:
    orange_field = orange_field_p[0]
    orange_field_descr = orange_field_p[1]
    ws.cell(q, 1).value = orange_field
    ws.cell(q, 2).value = orange_field_descr
    ws.cell(q, 1).font = Font(name='Cambria', size=12, bold=True)
    ws.cell(q, 2).font = Font(name='Cambria', size=11, bold=False, italic=True)
    ws.cell(q, 1).fill = PatternFill(start_color="FD6A02", end_color="FD6A02", fill_type="solid")
    ws.cell(q, 2).fill = PatternFill(start_color="FED8B1", end_color="FED8B1", fill_type="solid")
    ws.cell(q, 3).fill = PatternFill(start_color="FED8B1", end_color="FED8B1", fill_type="solid")
    q += 1
    ws.cell(q, 1).value = 'возм. значения'
    ws.cell(q, 2).value = 'краткое описание'
    ws.cell(q, 3).value = 'полное описание'
    ws.cell(q, 1).fill = PatternFill(start_color="FFFFE8", end_color="FFFFE8", fill_type="solid")
    ws.cell(q, 2).fill = PatternFill(start_color="FFFFE8", end_color="FFFFE8", fill_type="solid")
    ws.cell(q, 3).fill = PatternFill(start_color="FFFFE8", end_color="FFFFE8", fill_type="solid")
    # FFFFE8

    q += 1
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 55

    query3 = "SELECT " + orange_field + ", " + orange_field + "_rushort, " + orange_field + "_ru FROM tools.descr_" + orange_field
    res = get_db_rows(cu, query3)
    for row in res:
        ws.cell(q, 1).value = row[0]
        ws.cell(q, 2).value = row[1]
        ws.cell(q, 3).value = row[2]
        q += 1
    q += 1

src_wb = openpyxl.load_workbook(SM_SIMPLE_PATH)
src_sheet = src_wb.active
dso_file_previous.create_sheet("sm", 0)
dest_sheet = dso_file_previous.get_sheet_by_name("sm")

for i in range(1, src_sheet.max_row + 1):
    for j in range(1, src_sheet.max_column + 1):
        dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

dso_file_previous.save(dest_export_path)
cu.close()
co.close()
print('done')
